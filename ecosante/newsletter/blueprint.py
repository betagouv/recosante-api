from flask import (
    abort,
    jsonify,
    render_template,
    request,
    redirect,
    url_for,
    stream_with_context,
)
from flask.wrappers import Response
from datetime import datetime, timedelta
import io
from openpyxl import Workbook

from indice_pollution.helpers import today
from ecosante.inscription.models import Inscription

from ecosante.utils.decorators import admin_capability_url
from ecosante.utils import Blueprint
from .forms import FormAvis
from .models import Newsletter, NewsletterDB, db
from .tasks.import_in_sb import create_campaign, import_, send
from .tasks.send_webpush_notifications import send_webpush_notification, vapid_claims
from indice_pollution.history.models import IndiceATMO

bp = Blueprint("newsletter", __name__)

@bp.route('<short_id>/avis', methods=['GET', 'POST'])
def avis(short_id):
    nl = db.session.query(NewsletterDB).filter_by(short_id=short_id).first()
    if not nl:
        abort(404)
    nl.appliquee = request.args.get('avis') == 'oui' or request.args.get('appliquee') == 'oui'
    form = FormAvis(request.form or request.json, obj=nl)
    if request.method=='POST' and form.validate_on_submit():
        form.populate_obj(nl)
        db.session.add(nl)
        db.session.commit()
        if not request.accept_mimetypes.accept_html:
            return {
                "short_id": nl.short_id,
                "avis": nl.avis,
                "recommandation": nl.recommandation,
                "appliquee": nl.appliquee
            }
        return redirect(
            url_for('newsletter.avis_enregistre', short_id=short_id)
        )
    db.session.add(nl)
    db.session.commit()

    if not request.accept_mimetypes.accept_html:
        return {
            "short_id": nl.short_id,
            "avis": nl.avis,
            "recommandation": nl.recommandation,
            "appliquee": nl.appliquee
        }
    return render_template(
        'avis.html',
        nl=nl,
        form=form,
    )

@bp.route('<short_id>/avis/enregistre')
def avis_enregistre(short_id):
    return render_template('avis_enregistre.html')

@bp.route('<secret_slug>/avis/liste')
@bp.route('/avis/liste')
@admin_capability_url
def liste_avis():
    newsletters = NewsletterDB.query\
        .filter(NewsletterDB.avis.isnot(None))\
        .order_by(NewsletterDB.date.desc())\
        .all()
    return render_template('liste_avis.html', newsletters=newsletters)


@bp.route('<secret_slug>/avis/csv')
@bp.route('/avis/csv')
@admin_capability_url
def export_avis():
    return Response(
        stream_with_context(
            NewsletterDB.generate_csv_avis()
        ),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=export-avis-{datetime.now()}"
        }
    )

@bp.route('<secret_slug>/send_campaign/', methods=['GET', 'POST'])
@admin_capability_url
def send_campaign():
    now = request.args.get('now')
    mail_list_id = request.args.get('mail_list_id', type=int)
    campaign_id = create_campaign(now, mail_list_id)
    send(campaign_id)
    return "ok"

@bp.route('<secret_slug>/export/', methods=['GET', 'POST'])
@admin_capability_url
def export():
    mail_list_id = request.args.get('mail_list_id', type=int)
    if not mail_list_id:
        return "no mail list id", 404
    newsletters = NewsletterDB.query.filter_by(mail_list_id=mail_list_id)

    fieldnames = {k: i+1 for i, k in enumerate(newsletters[0].attributes().keys())}
    wb = Workbook()
    ws1 = wb.active
    for k, i in fieldnames.items():
        _ = ws1.cell(row=1, column=i, value=k)
    for nl in newsletters:
        for k, v in nl.attributes().items():
            _ = ws1.cell(row=i+2, column=fieldnames[k], value=v)
    output = io.BytesIO()
    wb.save(output)
    return output

@bp.route('<secret_slug>/test', methods=['GET', 'POST'])
@bp.route('/test', methods=['GET', 'POST'])
@admin_capability_url
def test():
    if request.method == "GET":
        return render_template("test.html")
    indice_atmo = int(request.form.get("indice_atmo"))
    uid = request.form.get("uid")
    inscription = Inscription.query.filter_by(uid=uid).first()
    nb_mails = 0
    nb_notifications = 0
    nb_notifications_sent = 0
    for media in inscription.indicateurs_media:
        nl = Newsletter(
            inscription=inscription,
            forecast={"data":[{"date": str(today()), "label": IndiceATMO.label_from_valeur(indice_atmo), "couleur": IndiceATMO.couleur_from_valeur(indice_atmo)}]},
            raep=int(request.form.get("raep")),
            allergenes={k: v for k, v in zip(request.form.getlist('allergene_nom'), request.form.getlist('allergene_value'))},
            validite_raep={
                "debut": today().strftime("%d/%m/%Y"),
                "fin": (today()+timedelta(days=7)).strftime("%d/%m/%Y")
            }
        )
        if media == "mail":
            result = import_(None, [NewsletterDB(nl)], force_send=True, test=True)
            send(None, result["email_campaign_id"], test=True)
            nb_mails += 1
        elif media == "notifications_web":
            for wp in inscription.webpush_subscriptions_info:
                nl.webpush_subscription_info = wp
                nl.webpush_subscription_info_id = wp.id
                if send_webpush_notification(NewsletterDB(nl), vapid_claims):
                    nb_notifications_sent += 1
                nb_notifications += 1

    return render_template("test_ok.html", nb_mails=nb_mails, nb_notifications=nb_notifications, nb_notifications_sent=nb_notifications_sent)